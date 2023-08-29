import tkinter as tk
from tkinter import filedialog
import os
import subprocess
import sys
import ctypes
import re
import logging

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
except Exception as e:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

try:
    from tqdm import tqdm
except Exception as e:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm"])

root = tk.Tk() # -------------------------------------------------------------------- Dialog print shit
root.withdraw()

class standard:
    Code = {}
    wb = load_workbook(filename = filedialog.askopenfilename(filetypes=[("Conversion Standard", "*.xlsx")]))
    ws = wb.active
    for i in range(ws.max_row):
        SHA_Code = ws.cell(row=i+1, column=6).value
        SHA_Attrib = ws.cell(row=i+1, column=7).value
        TO_Code = ws.cell(row=i+1, column=5).value
        if ws.cell(row=i+1, column=7).value is not None:
            SHA_Dict = SHA_Code + " " + SHA_Attrib
        else:
            SHA_Dict = SHA_Code
        Code[SHA_Dict] = TO_Code

# ---- #

def loadFile(filename): # ----------------------------------------------------------- Loading all data of file into 'fileList'
    with open(filename, 'r') as file:
        fileList = []
        for count, line in enumerate(file): # --------------------------------------- It's gross but it works. Better than dealing with XML
            fileList.append(line)
    return fileList # --------------------------------------------------------------- Return file as list.

# ---- #

def checkNums(string):
    try:
        int(string)
        return False
    except Exception as e:
        pass
    return any(char.isdigit() for char in string)

def splitNums(input):
    match = re.match(r"([a-z]+)([0-9]+)", input, re.I)
    if match:
        items = match.groups()
    return items

def convert(file, line, type):
    location = file.index(line)
    if type:
        code = file[location+2]
        featureLook = location+7
    else:
        code = file[location+3]
        featureLook = location+6
    
    if "<Source>" in file[featureLook]:
        featureLook = featureLook+1
    inputCode = code.replace("<Code>", "").replace("</Code>",  "").strip()
    inputCode = inputCode.split(" ")
    
    baseCode = []
    featureCodes = []
    for code in inputCode:
        if checkNums(code):
            numCode = splitNums(code)
            baseCode.append(numCode[0])
            baseCode.append(numCode[1])
        else:
            baseCode.append(code)

    if "<Features>" in file[featureLook]:
        featureExtract = []

        count = 0
        while True:
            currentLine = file[featureLook+count]
            featureExtract.append(currentLine)
            if "</Features>" in currentLine:
                break
            else:
                count += 1

        for num, line in enumerate(featureExtract):
            if "<Name>" in line:
                Attrib = file[featureLook+num+1]
                Attrib = Attrib.replace("<Value>", "").replace("</Value>",  "").strip()
                if Attrib == "<Value/>":
                    pass
                else:
                    featureCodes = Attrib.split(' ')

    convertedCode = []
    flag = 0
    for code in baseCode:
        for feat in featureCodes:
            try:
                convert = standard.Code[code + " " + feat]
                flag = 1
                break
            except Exception as e:
                pass
        if flag == 1:
            flag = 0
        else:
            try:
                convert = standard.Code[code]
            except Exception as e:
                convert = code
                logging.warning("Unknown Code! : {}".format(convert))

        convertedCode.append(convert)


    preCombined = ' '.join(convertedCode)

    

    if checkNums(preCombined):
        for count, item in enumerate(convertedCode):
            try:
                int(item)
                convertedCode[count-1] = convertedCode[count-1] + str(item)
                convertedCode.pop(count)
            except Exception as e:
                pass
            
    newCode = ' '.join(convertedCode)

    if type:
        code = file[location+2]
    else:
        code = file[location+3]
    codeSplit = code.split('<Code>')
    junk = codeSplit[1].split('</Code>')
    codeFormat = [codeSplit[0], '<Code>', newCode, '</Code>', junk[1]]

    return ''.join(codeFormat) # ---------------------------------------------------- Must return string properly formatted, easier to condence a list.

# ---- #

def export(file, filename): # ----------------------------------------------------------------- Takes in a list as the file
    f = open(filename[:-4] + "_Converted.jxl" , "w")
    for line in file:
        f.write(line) # ------------------------------------------------------------- Writes each item in list as a line.
    f.close()
    ctypes.windll.user32.MessageBoxW(0, "Finished!", "Converter", 0)

# ---- #

def main():
    filename = filedialog.askopenfilename(filetypes=[("JXL File", "*.jxl")]) # ------ Dialog prompt for finding JXL file
    logging.basicConfig(filename=filename[:-4] + ".log", filemode='w', format='%(levelname)s - %(message)s')
    logging.debug('Starting')
    file = loadFile(filename) # ----------------------------------------------------- Calls load file

    newFile = []
    flag = 0 # ---------------------------------------------------------------------- Used for skipping a set amount of lines for proper formatting
    os.system('cls')
    for num, line in enumerate(tqdm(file)):
        
        if flag != 0: # ------------------------------------------------------------- Method for skipping a line
            flag -= 1
        
        elif "PointRecord " in line: # ---------------------------------------------- Looking for 'PointRecorded', ignoring '/PointRecorded'
            location = file.index(line) 
            toCode = file[location+2]
            fromCode = convert(file, line, True) # -------------------------------------------- Calling convert, should return formatted string

            newFile.append(line)
            newFile.append(file[location+1]) # -------------------------------------- Writting lines prior to formatted code instead of trying to-
            newFile.append(fromCode) # ---------------------------------------------- break out of the 'for line in file' loop

            flag = 2 # -------------------------------------------------------------- Skipping lines already written

        elif "<Point>" in line: # --------------------------------------------------- Looking for '<Point>', ignoring '</Point>'
            location = file.index(line)
            toCode = file[num+3]
            fromCode = convert(file, line, False) # -------------------------------------------- Same as prior block

            newFile.append(line)
            newFile.append(file[num+1])
            newFile.append(file[num+2])
            newFile.append(fromCode)

            flag = 3 # -------------------------------------------------------------- Skipping lines already written
        
        else:
            newFile.append(line) # -------------------------------------------------- Lines to not be converted are skipped and written as normal
    
    export(newFile, filename) # --------------------------------------------------------------- Calls export with file as list, filename, and standard.

# ---- #

if __name__ == "__main__": # -------------------------------------------------------- Script initalizing
    main()
    logging.debug('Finished')